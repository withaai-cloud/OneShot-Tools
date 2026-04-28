from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
import os
import sys
import re
import csv
import base64
import json
import urllib.request
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import zipfile
from io import BytesIO, StringIO

app = Flask(__name__, template_folder='api/templates', static_folder='api/static')

UPLOAD_FOLDER = '/tmp/uploads'
OUTPUT_FOLDER = '/tmp/outputs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# ─────────────────────────────────────────────
# ABSA CHARACTER DECODING (for old-style PDFs)
# ─────────────────────────────────────────────

ABSA_CHAR_MAP = {
    # Digits
    'ð': '0', 'ñ': '1', 'ò': '2', 'ó': '3', 'ô': '4',
    'õ': '5', 'ö': '6', '÷': '7', 'ø': '8', 'ù': '9',
    # Punctuation/special
    'a': '/', 'k': '.', 'K': '.', '@': ' ', '`': '-', 'z': ':',
    '\\': '*', 'm': 'j',
    # Control chars
    '\x81': ' ', '\x82': 'b', '\x83': 'c', '\x84': 'o', '\x85': 'l',
    '\x86': 'i', '\x87': 'g', '\x88': 'h', '\x89': 'n', '\x8a': 'j',
    '\x8b': 'k', '\x8c': 'w', '\x8d': 'd', '\x8e': 'f', '\x8f': 'p',
    '\x92': 'a', '\x95': 'l', '\x99': 'r',
    # Uppercase
    'Á': 'A', 'Â': 'B', 'Ã': 'C', 'Ä': 'D', 'Å': 'E', 'Æ': 'F',
    'Ç': 'G', 'È': 'H', 'É': 'I', 'Ê': 'J', 'Ë': 'K', 'Ì': 'L',
    'Í': 'M', 'Î': 'N', 'Ï': 'O', 'Ñ': 'P', 'Ò': 'Q', 'Ó': 'R',
    'Ô': 'S', 'Õ': 'T', 'Ö': 'U', '×': 'V', 'Ø': 'W', 'Ù': 'X',
    'Ú': 'Y', 'Û': 'Z',
    # Lowercase
    '¢': 'e', '£': 't', '¤': 'a', '¥': 'r', '¦': 'w', '§': 'y',
    '¨': 'n', '©': 'o', 'ª': 'i', '«': 'u', '¬': 's', '\xad': 'd',
    '®': 'l', '¯': 'c', '°': 'f', '±': 'h', '²': 'm', '³': 'p',
    '´': 'g', 'µ': 'b', '¶': 'v', '·': 'k', '¸': 'x', '¹': 'j',
    'º': 'q', '»': 'z',
}

ABSA_WORD_CORRECTIONS = {
    'Abs/': 'Absa', 'B/n': 'Bank', 'B/nk': 'Bank',
    'P/rtners': 'Partners', 'P/rtner': 'Partner',
    'Ab/qulusi': 'Abaqulusi', 'S/rsef': 'Sarsef',
    'C/sh': 'Cash', 'Petty C/sh': 'Petty Cash',
    'P/r.ing': 'Parking', 'P/r.i': 'Parki',
    'G/s': 'Gas', 'Zulu G/s': 'Zulu Gas',
    'M/r.r/j': 'Markray', 'Pr M/r.r/j': 'Pr Markray',
    'Contr/.tors': 'Contractors', 'Ms Contr/ctors': 'Ms Contractors',
    'C .usel': 'Cousel', 'Bl/dsy': 'Bladsy',
    'v/n': 'van', 'D/tuj': 'Datum',
    'Tr/ns/.sies': 'Transaksies', 'Tr/ns/.siebes.rywing': 'Transaksiebeskrywing',
    'Bes.rywing': 'Beskrywing', '.oste': 'Koste',
    'Debietbedr/g': 'Debietbedrag', '.redietbedr/g': 'Kredietbedrag',
    'S/ldo': 'Saldo', 'Bl/dsy': 'Bladsy',
    'Registr/sie': 'Registrasie', 'Bel/stingf/.tuur': 'Belastingfaktuur',
    'Fin/nsiële': 'Finansiële', 'dienstevers./ffer': 'diensteverskkaffer',
    '.redietvers./ffer': 'Kredietverskkaffer', 'Gej/gtigde': 'Gemagtigde',
    'Beper.': 'Beperk', 'nojjer': 'nommer',
    'Tje.re.ening': 'Tjekrekening', '.w/': 'Kwa',
    'Pro Gu/rd': 'Pro Guard', '.9 Pro': 'K9 Pro',
    'Psgojic': 'Psgkons', 'Psg.ons': 'Psgkons',
    'Acb Gerd/': 'Acb Gerda', 'Gerd/': 'Gerda',
    'Adjin': 'Admin', 'ADJIN': 'ADMIN',
    'Mndeli': 'Mndelik', 'MNDL.S': 'MNDLKS',
    'RE.-FOOI': 'REK-FOOI', '.REDIETRENTE': 'KREDIETRENTE',
    '.ONTANTDEPOSITO': 'KONTANTDEPOSITO', 'TRANSA.SIE': 'TRANSAKSIE',
    '.ONTA.': 'KONTAK', 'BESIGHEIDSBAN.': 'BESIGHEIDBANK',
    'ABSA.CO.ZA': 'ABSA.CO.ZA',
}

def decode_absa_text(text):
    """Decode garbled ABSA PDF text using character map"""
    result = ''
    for char in text:
        result += ABSA_CHAR_MAP.get(char, char)
    return result

def apply_word_corrections(text):
    """Apply known word-level corrections to ABSA descriptions"""
    for wrong, right in ABSA_WORD_CORRECTIONS.items():
        text = text.replace(wrong, right)
    return text

# ─────────────────────────────────────────────
# PDF TYPE DETECTION
# ─────────────────────────────────────────────

def is_image_based_pdf(pdf_path):
    """Return True if PDF pages contain images only (no text layer)"""
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            text = page.get_text()
            if text.strip():
                return False
        return True
    except Exception:
        return True

def get_page_images_b64(pdf_path):
    """Extract each page as a base64 JPEG from an image-based PDF"""
    doc = fitz.open(pdf_path)
    images = []
    for page in doc:
        d = page.get_text('rawdict')
        blocks = d.get('blocks', [])
        if blocks and blocks[0].get('type') == 1:
            # Embedded JPEG/image
            img_bytes = blocks[0]['image']
            images.append(base64.b64encode(img_bytes).decode())
        else:
            # Render page as image if no embedded image found
            mat = fitz.Matrix(2, 2)
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes('jpeg')
            images.append(base64.b64encode(img_bytes).decode())
    return images

# ─────────────────────────────────────────────
# VISION-BASED EXTRACTION (Claude API)
# ─────────────────────────────────────────────

VISION_SYSTEM_PROMPT = """You are a bank statement parser. Extract transactions from ABSA Tjekrekeningstaat (bank statement) images.

Return ONLY a JSON array — no markdown, no explanation, no preamble.

For each transaction row in the "U transaksies" table:
- date: string in DD/MM/YYYY format
- description: full description (merge main line + sub-line if present, separated by space)
- amount: number (negative for debits/Debietbedrag, positive for credits/Kredietbedrag)

The table columns are: Datum | Transaksiebeskrywing | Koste | [type] | Debietbedrag | Kredietbedrag | Saldo

Numbers use period as decimal separator. Ignore "Saldo Oorgedra" (opening balance) rows.

Example output:
[{"date":"05/11/2025","description":"Digitale Betaal Dt Vereffenin Absa Bank Pr Markram Musi","amount":-920.00},{"date":"07/11/2025","description":"Digitale Betaal Kt Vereffenin Absa Bank Uys & Partners Huur","amount":37443.45}]"""

def extract_via_vision(pdf_path):
    """Use Claude vision API to extract transactions from image-based PDFs"""
    images_b64 = get_page_images_b64(pdf_path)
    all_transactions = []

    for img_b64 in images_b64:
        content = [
            {
                "type": "image",
                "source": {"type": "base64", "media_type": "image/jpeg", "data": img_b64}
            },
            {
                "type": "text",
                "text": "Extract all transaction rows from this bank statement page. Return JSON array only."
            }
        ]

        payload = json.dumps({
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 4000,
            "system": VISION_SYSTEM_PROMPT,
            "messages": [{"role": "user", "content": content}]
        }).encode()

        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY environment variable is not set. Add it in Vercel project settings.")

        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'anthropic-version': '2023-06-01',
                'x-api-key': api_key
            }
        )

        try:
            with urllib.request.urlopen(req, timeout=90) as resp:
                raw = resp.read()
                result = json.loads(raw)
                print(f"[VISION] API response keys: {list(result.keys())}", file=sys.stderr)
                if 'error' in result:
                    print(f"[VISION] API error: {result['error']}", file=sys.stderr)
                    continue
                text = result['content'][0]['text'].strip()
                print(f"[VISION] Raw text ({len(text)} chars): {text[:300]}", file=sys.stderr)
                # Strip any markdown fences if present
                text = re.sub(r'^```(?:json)?\s*', '', text)
                text = re.sub(r'\s*```$', '', text)
                rows = json.loads(text)
                print(f"[VISION] Parsed {len(rows)} rows", file=sys.stderr)
                all_transactions.extend(rows)
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            print(f"[VISION] HTTP {e.code} error: {body}", file=sys.stderr)
            continue
        except urllib.error.URLError as e:
            print(f"[VISION] URL error: {e.reason}", file=sys.stderr)
            continue
        except json.JSONDecodeError as e:
            print(f"[VISION] JSON parse error: {e} — text was: {text[:200]}", file=sys.stderr)
            continue
        except Exception as e:
            import traceback
            print(f"[VISION] Unexpected error: {traceback.format_exc()}", file=sys.stderr)
            continue

    # Convert to standard internal format: (date_str, description, amount_float)
    transactions = []
    for row in all_transactions:
        try:
            date_str = row.get('date', '').strip()
            description = row.get('description', '').strip()
            amount = float(row.get('amount', 0))
            if date_str and description:
                transactions.append((date_str, description, amount))
        except (ValueError, TypeError):
            continue

    return transactions

# ─────────────────────────────────────────────
# TEXT-BASED EXTRACTION (FNB, Standard Bank, old ABSA)
# ─────────────────────────────────────────────

def detect_bank(text):
    """Detect bank from PDF text"""
    text_lower = text.lower()
    if 'absa' in text_lower or 'tjekrekeningstaat' in text_lower:
        return 'ABSA'
    elif 'standard bank' in text_lower or 'standardbank' in text_lower:
        return 'STANDARD'
    elif 'fnb' in text_lower or 'first national bank' in text_lower:
        return 'FNB'
    return 'FNB'

def extract_fnb_transactions(text):
    """Extract transactions from FNB bank statement text"""
    transactions = []
    lines = text.split('\n')
    date_pattern = re.compile(r'^\d{1,2}\s+\w+\s+\d{2,4}')
    amount_pattern = re.compile(r'[-+]?\d{1,3}(?:,\d{3})*\.\d{2}')

    for line in lines:
        line = line.strip()
        if not line:
            continue
        date_match = date_pattern.match(line)
        if date_match:
            amounts = amount_pattern.findall(line)
            if amounts:
                date_str = parse_fnb_date(date_match.group())
                description_start = date_match.end()
                description = line[description_start:].strip()
                for amt in amounts:
                    description = description.replace(amt, '').strip()
                amount_str = amounts[0].replace(',', '')
                try:
                    amount = float(amount_str)
                    transactions.append((date_str, description, amount))
                except ValueError:
                    pass
    return transactions

def parse_fnb_date(date_str):
    """Parse FNB date format like '30 Oct 25' or '30 Oct 2025'"""
    months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
        'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
        'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    parts = date_str.split()
    if len(parts) >= 3:
        day = parts[0].zfill(2)
        month = months.get(parts[1], '01')
        year = parts[2]
        if len(year) == 2:
            year_int = int(year)
            year = f'20{year}' if year_int <= 50 else f'19{year}'
        return f'{day}/{month}/{year}'
    return date_str

def extract_absa_transactions_text(text):
    """Extract transactions from old-style ABSA PDF text (encoded chars)"""
    decoded = decode_absa_text(text)
    transactions = []
    lines = decoded.split('\n')
    date_pattern = re.compile(r'^\d{2}/\d{2}/\d{4}')
    amount_pattern = re.compile(r'\d{1,3}(?:,\d{3})*\.\d{2}')

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        date_match = date_pattern.match(line)
        if date_match:
            date_str = line[:10]
            rest = line[10:].strip()
            # Check next line for continuation
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line and not date_pattern.match(next_line):
                    rest = rest + ' ' + next_line
                    i += 1
            amounts = amount_pattern.findall(rest)
            if amounts:
                description = rest
                for amt in amounts:
                    description = description.replace(amt, '').strip()
                description = apply_word_corrections(description)
                description = re.sub(r'\s+', ' ', description).strip()
                amount_val = amounts[-1].replace(',', '')
                try:
                    amount = float(amount_val)
                    # If there's a debit indicator, make negative
                    if re.search(r'\bD\b|\bdebit\b', rest, re.IGNORECASE):
                        amount = -abs(amount)
                    transactions.append((date_str, description, amount))
                except ValueError:
                    pass
        i += 1
    return transactions

def extract_standard_bank_transactions(text):
    """Extract transactions from Standard Bank statement text"""
    transactions = []
    lines = text.split('\n')
    date_pattern = re.compile(r'^\d{2}\s+\w{3}\s+\d{2}')
    amount_pattern = re.compile(r'[-+]?\d{1,3}(?:,\d{3})*\.\d{2}')

    for line in lines:
        line = line.strip()
        if not line:
            continue
        date_match = date_pattern.match(line)
        if date_match:
            amounts = amount_pattern.findall(line)
            if amounts:
                date_str = parse_fnb_date(date_match.group())
                description_start = date_match.end()
                description = line[description_start:].strip()
                for amt in amounts:
                    description = description.replace(amt, '').strip()
                amount_str = amounts[0].replace(',', '')
                try:
                    amount = float(amount_str)
                    transactions.append((date_str, description, amount))
                except ValueError:
                    pass
    return transactions

def extract_transactions_from_pdf(filepath, invert_amounts=False):
    """Main extraction function — detects PDF type and uses correct method"""

    # ── Image-based PDF: use Claude vision ──
    if is_image_based_pdf(filepath):
        print(f"[INFO] Image-based PDF detected: {filepath}", file=sys.stderr)
        transactions = extract_via_vision(filepath)
        if invert_amounts:
            transactions = [(d, desc, -amt) for d, desc, amt in transactions]
        return transactions

    # ── Text-based PDF: use pypdfium2 or fitz ──
    try:
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(filepath)
        full_text = ''
        for page in pdf:
            textpage = page.get_textpage()
            full_text += textpage.get_text_range() + '\n'
    except Exception:
        doc = fitz.open(filepath)
        full_text = ''
        for page in doc:
            full_text += page.get_text() + '\n'

    bank = detect_bank(full_text)
    print(f"[INFO] Detected bank: {bank}", file=sys.stderr)

    if bank == 'ABSA':
        transactions = extract_absa_transactions_text(full_text)
    elif bank == 'STANDARD':
        transactions = extract_standard_bank_transactions(full_text)
    else:
        transactions = extract_fnb_transactions(full_text)

    if invert_amounts:
        transactions = [(d, desc, -amt) for d, desc, amt in transactions]

    return transactions

# ─────────────────────────────────────────────
# OUTPUT FILE CREATION
# ─────────────────────────────────────────────

def create_excel_file(transactions, output_path):
    """Create Excel file from transactions list"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ['Date', 'Description', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_idx, (date_str, description, amount) in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=date_str)
        ws.cell(row=row_idx, column=2, value=description)
        amount_cell = ws.cell(row=row_idx, column=3, value=amount)
        if amount < 0:
            amount_cell.font = Font(color="C00000")

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15

    wb.save(output_path)

def create_csv_file(transactions, output_path):
    """Create CSV file from transactions list"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Date', 'Description', 'Amount'])
        for date_str, description, amount in transactions:
            writer.writerow([date_str, description, amount])

# ─────────────────────────────────────────────
# FLASK ROUTES
# ─────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/tools/bank-statement-converter')
def bank_statement_converter():
    return render_template('bank_statement_converter.html')

@app.route('/tools/tax-optimizer')
def tax_optimizer():
    return render_template('tax_optimizer.html')

@app.route('/convert', methods=['POST'])
def convert():
    """Handle PDF conversion"""
    try:
        if 'files[]' not in request.files:
            return jsonify({'error': 'No files uploaded'}), 400

        files = request.files.getlist('files[]')
        if not files or files[0].filename == '':
            return jsonify({'error': 'No files selected'}), 400

        invert_amounts = request.form.get('invert_amounts') == 'true'
        output_format = request.form.get('output_format', 'xlsx')  # 'xlsx' or 'csv'

        output_files = []
        all_transactions = []

        for file in files:
            if file and file.filename.endswith('.pdf'):
                filename = file.filename
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                transactions = extract_transactions_from_pdf(filepath, invert_amounts)
                all_transactions.extend(transactions)

                if output_format == 'csv':
                    output_filename = filename.replace('.pdf', '_transactions.csv')
                    output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
                    create_csv_file(transactions, output_path)
                else:
                    output_filename = filename.replace('.pdf', '_transactions.xlsx')
                    output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
                    create_excel_file(transactions, output_path)

                # Read file bytes for response
                with open(output_path, 'rb') as f:
                    file_bytes = f.read()

                output_files.append({
                    'original': filename,
                    'output': output_filename,
                    'transactions': len(transactions),
                    'file_data': base64.b64encode(file_bytes).decode()
                })

                os.remove(filepath)
                os.remove(output_path)

        # Multiple files: also create combined
        if len(output_files) > 1:
            if output_format == 'csv':
                combined_name = f'combined_transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                combined_path = os.path.join(app.config['OUTPUT_FOLDER'], combined_name)
                create_csv_file(all_transactions, combined_path)
            else:
                combined_name = f'combined_transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                combined_path = os.path.join(app.config['OUTPUT_FOLDER'], combined_name)
                create_excel_file(all_transactions, combined_path)

            with open(combined_path, 'rb') as f:
                combined_bytes = f.read()
            os.remove(combined_path)

            return jsonify({
                'success': True,
                'multiple': True,
                'files': output_files,
                'combined_file': combined_name,
                'combined_data': base64.b64encode(combined_bytes).decode(),
                'total_transactions': len(all_transactions)
            })
        else:
            return jsonify({
                'success': True,
                'multiple': False,
                'file': output_files[0]['output'],
                'transactions': output_files[0]['transactions'],
                'file_data': output_files[0]['file_data']
            })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/test-vision', methods=['POST'])
def test_vision():
    """Test endpoint - upload a PDF and see exactly what happens"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'})
    
    file = request.files['file']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)
    
    result = {'filename': file.filename, 'steps': []}
    
    # Step 1: Check if image-based
    try:
        doc = fitz.open(filepath)
        page_texts = []
        for i, page in enumerate(doc):
            t = page.get_text()
            page_texts.append(f"page{i}: {len(t)} chars")
        result['steps'].append(f"fitz opened OK, pages: {page_texts}")
        img_based = is_image_based_pdf(filepath)
        result['steps'].append(f"is_image_based: {img_based}")
    except Exception as e:
        result['steps'].append(f"fitz ERROR: {e}")
        os.remove(filepath)
        return jsonify(result)
    
    if not img_based:
        result['steps'].append("NOT image-based - would use text extraction")
        os.remove(filepath)
        return jsonify(result)
    
    # Step 2: Extract images
    try:
        imgs = get_page_images_b64(filepath)
        result['steps'].append(f"extracted {len(imgs)} page images, sizes: {[len(i) for i in imgs]}")
    except Exception as e:
        result['steps'].append(f"image extraction ERROR: {e}")
        os.remove(filepath)
        return jsonify(result)
    
    # Step 3: Call vision API on first page only
    try:
        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        content_msg = [
            {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": imgs[0]}},
            {"type": "text", "text": "Extract all transaction rows from this bank statement page. Return JSON array only."}
        ]
        payload = json.dumps({
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 4000,
            "system": VISION_SYSTEM_PROMPT,
            "messages": [{"role": "user", "content": content_msg}]
        }).encode()
        
        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'anthropic-version': '2023-06-01',
                'x-api-key': api_key
            }
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            api_result = json.loads(resp.read())
            raw_text = api_result['content'][0]['text']
            result['steps'].append(f"API call OK, raw response ({len(raw_text)} chars): {raw_text[:500]}")
    except Exception as e:
        result['steps'].append(f"API call ERROR: {type(e).__name__}: {e}")
    
    os.remove(filepath)
    return jsonify(result)


@app.route('/debug')
def debug():
    """Debug endpoint to check environment"""
    import sys
    info = {
        'python': sys.version,
        'env_has_api_key': bool(os.environ.get('ANTHROPIC_API_KEY')),
        'api_key_prefix': os.environ.get('ANTHROPIC_API_KEY', '')[:8] + '...' if os.environ.get('ANTHROPIC_API_KEY') else 'NOT SET',
    }
    # Check libraries
    try:
        import fitz
        info['pymupdf'] = fitz.__version__
    except ImportError as e:
        info['pymupdf'] = f'MISSING: {e}'
    try:
        import pypdfium2
        info['pypdfium2'] = 'available'
    except ImportError:
        info['pypdfium2'] = 'not installed'
    try:
        import openpyxl
        info['openpyxl'] = openpyxl.__version__
    except ImportError as e:
        info['openpyxl'] = f'MISSING: {e}'
    return jsonify(info)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
